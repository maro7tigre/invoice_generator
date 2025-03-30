import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText
from copy import copy

class InvoiceGenerator:
    def __init__(self):
        # Determine OS and set appropriate paths
        self.is_windows = os.name == 'nt'
        
        # Default paths - handles both Windows and Linux paths
        if self.is_windows:
            # Windows paths
            self.base_path = str(Path.home() / "Desktop" / "facture")
        else:
            # Linux paths
            self.base_path = str(Path.home() / "Desktop" / "facture")
            
        # Make sure the base directory exists
        os.makedirs(self.base_path, exist_ok=True)
            
        self.template_path = os.path.join(self.base_path, "FACTURE COMPT.xlsx")
        self.output_folder = self.base_path
    
    def load_data(self, data_file):
        """Load data from Excel file and extract description, quantity, and unit price"""
        try:
            # Read the first sheet of the Excel file
            df = pd.read_excel(data_file, header=None)
            
            # Check if the dataframe has the expected structure
            if df.shape[1] < 3:
                raise ValueError("Fichier incorrect: colonnes insuffisantes")
            
            # Extract line items (Description, Quantity, Unit Price)
            line_items = []
            for idx, row in df.iterrows():
                # Skip rows where either description, quantity or price is missing
                if pd.notna(row[0]) and pd.notna(row[1]) and pd.notna(row[2]):
                    try:
                        # Convert to appropriate types
                        description = str(row[0])
                        quantity = float(row[1])
                        unit_price = float(row[2])
                        
                        line_items.append({
                            'description': description,
                            'quantity': quantity,
                            'unit_price': unit_price
                        })
                    except (ValueError, TypeError):
                        # Skip rows where conversion to float fails
                        print(f"Skipping row {idx+1}: Could not convert quantity or unit price to number")
                        continue
            
            # Return only the line items - calculations will be handled by the template
            return {
                'line_items': line_items
            }
        except Exception as e:
            print(f"Error loading data: {str(e)}")
            raise
    
    def generate_invoice_id(self):
        """Generate a unique invoice ID in the format FA XXX/YYYY"""
        # Get current year
        year = datetime.now().year
        
        # Look for existing invoice files in the output directory to determine the next number
        invoice_files = [f for f in os.listdir(self.output_folder) 
                        if f.startswith('invoice_') and f.endswith('.xlsx')]
        
        if not invoice_files:
            # No existing files, start with 001
            next_num = 1
        else:
            # Extract numbers from existing files and find the max
            nums = []
            for f in invoice_files:
                try:
                    # Extract the number part from filenames like 'invoice_001.xlsx'
                    num_part = f.replace('invoice_', '').replace('.xlsx', '')
                    # Handle cases where there might be _1, _2 suffixes for multiple invoices
                    if '_' in num_part:
                        num_part = num_part.split('_')[0]
                    nums.append(int(num_part))
                except ValueError:
                    continue
            
            next_num = max(nums) + 1 if nums else 1
        
        # Format as 3-digit number
        formatted_num = f"{next_num:03d}"
        
        # Return both the file ID and the display ID
        return formatted_num, f"FA {formatted_num}/{year}"
    
    def create_invoice(self, data_file, invoice_id=None, output_file=None, client_info=None):
    
        try:
            # Check if template exists
            if not os.path.exists(self.template_path):
                raise FileNotFoundError(f"Le fichier modèle n'existe pas: {self.template_path}")
                    
            # Check if data file exists
            if not os.path.exists(data_file):
                raise FileNotFoundError(f"Le fichier de données n'existe pas: {data_file}")
            
            # Load data without auto-calculating
            data = self.load_data(data_file)
            
            # Get total number of items
            total_items = len(data['line_items'])
            
            # Set max rows per invoice to match your template
            max_rows_per_invoice = 23  # Adjusted to match your 23-row template
            
            # Calculate how many invoices we need
            num_invoices = (total_items + max_rows_per_invoice - 1) // max_rows_per_invoice
            
            # Store all generated invoice paths
            generated_invoices = []
            
            # Generate invoice ID if not provided
            initial_file_id, initial_display_id = self.generate_invoice_id() if invoice_id is None else (invoice_id, f"FA {invoice_id}/{datetime.now().year}")
            file_id = initial_file_id
            
            # Define start_row here so it's available in all scopes
            start_row = 12  # First row of items
            
            # Process each batch of items
            for invoice_index in range(num_invoices):
                # Calculate the start and end indices for this invoice
                start_idx = invoice_index * max_rows_per_invoice
                end_idx = min((invoice_index + 1) * max_rows_per_invoice, total_items)
                
                # Get items for this invoice
                invoice_items = data['line_items'][start_idx:end_idx]
                
                # Set the output filename for this invoice
                if output_file is None:
                    # If multiple invoices, append a suffix to the filename
                    if num_invoices > 1:
                        current_file_id = f"{file_id}_{invoice_index + 1}"
                        current_display_id = f"{initial_display_id}_{invoice_index + 1}"
                    else:
                        current_file_id = file_id
                        current_display_id = initial_display_id
                    
                    current_output_file = os.path.join(self.output_folder, f"invoice_{current_file_id}.xlsx")
                else:
                    # If output_file was specified but we have multiple invoices, add suffix
                    if num_invoices > 1:
                        base, ext = os.path.splitext(output_file)
                        current_output_file = f"{base}_{invoice_index + 1}{ext}"
                        current_display_id = f"{initial_display_id}_{invoice_index + 1}"
                    else:
                        current_output_file = output_file
                        current_display_id = initial_display_id
                
                # Make sure output directory exists
                os.makedirs(os.path.dirname(current_output_file), exist_ok=True)
                
                # Copy the template file directly - safest approach
                import shutil
                shutil.copy2(self.template_path, current_output_file)
                
                # Open the workbook and try a more direct approach
                import openpyxl
                workbook = openpyxl.load_workbook(current_output_file)
                sheet = workbook.active
                
                # Define direct cell references instead of trying to access merged cells
                
                # Set invoice number and date - direct references
                if 'E3' in sheet:
                    sheet['E3'].value = current_display_id
                
                if 'I3' in sheet:
                    sheet['I3'].value = datetime.now().strftime("%d/%m/%Y")
                
                # Set client info if provided - direct references
                if client_info:
                    if 'name' in client_info and 'H5' in sheet:
                        sheet['H5'].value = client_info['name']
                    if 'address' in client_info and 'H7' in sheet:
                        sheet['H7'].value = client_info['address']
                    if 'ice' in client_info and 'H9' in sheet:
                        sheet['H9'].value = client_info['ice']
                
                # Fill in items using explicit cell references
                for idx, item in enumerate(invoice_items):
                    row = start_row + idx
                    
                    # Description - try both A and B columns (since it might be a merged range)
                    try:
                        sheet.cell(row=row, column=1).value = item['description']
                    except:
                        try:
                            # Try the first cell in another column in case it's part of the merge
                            sheet.cell(row=row, column=2).value = item['description']
                        except:
                            print(f"Could not set description for row {row}")
                    
                    # Quantity (column H)
                    try:
                        sheet.cell(row=row, column=8).value = round(item['quantity'], 2)
                    except:
                        print(f"Could not set quantity for row {row}")
                        
                    # Unit price (column I)
                    try:
                        sheet.cell(row=row, column=9).value = round(item['unit_price'], 2)
                    except:
                        print(f"Could not set unit price for row {row}")
                    
                    # Calculate total for this row (quantity * unit price) - column J
                    try:
                        cell = sheet.cell(row=row, column=10)
                        cell.value = f"=H{row}*I{row}"
                    except:
                        print(f"Could not set total formula for row {row}")
                
                # Clear unnecessary zeros in rows 35 and 36 column J 
                try:
                    # Clear row 35 if it has a zero
                    cell_35 = sheet.cell(row=35, column=10)
                    if cell_35.value == 0 or cell_35.value == "0":
                        cell_35.value = None
                    
                    # Clear row 36 if it has a zero  
                    cell_36 = sheet.cell(row=36, column=10)
                    if cell_36.value == 0 or cell_36.value == "0":
                        cell_36.value = None
                except:
                    print("Could not clear zero values in rows 35/36")
                
                # Find the proper total cells - these should be in a separate table below
                # Identify them by checking columns G, H, or I for labels like "Total HT", "TVA", etc.
                total_ht_row = None
                tva_row = None
                total_ttc_row = None
                
                # Search for the total cells by looking for their labels
                for row in range(37, 45):  # Check rows after the main table
                    for col in range(7, 10):  # Check columns G, H, I
                        cell_value = sheet.cell(row=row, column=col).value
                        if cell_value and isinstance(cell_value, str):
                            cell_text = cell_value.lower()
                            if "total ht" in cell_text:
                                total_ht_row = row
                            elif "tva" in cell_text:
                                tva_row = row
                            elif "ttc" in cell_text or "total ttc" in cell_text:
                                total_ttc_row = row
                
                # If we found the total cells, update their formulas
                if total_ht_row:
                    try:
                        subtotal_range = f"J{start_row}:J{start_row + len(invoice_items) - 1}"
                        sheet.cell(row=total_ht_row, column=10).value = f"=SUM({subtotal_range})"
                        print(f"Set Total HT formula in row {total_ht_row}")
                    except:
                        print(f"Could not set Total HT formula")
                
                if tva_row and total_ht_row:
                    try:
                        sheet.cell(row=tva_row, column=10).value = f"=J{total_ht_row}*0.2"
                        print(f"Set TVA formula in row {tva_row}")
                    except:
                        print(f"Could not set TVA formula")
                
                if total_ttc_row and total_ht_row and tva_row:
                    try:
                        sheet.cell(row=total_ttc_row, column=10).value = f"=J{total_ht_row}+J{tva_row}"
                        print(f"Set Total TTC formula in row {total_ttc_row}")
                    except:
                        print(f"Could not set Total TTC formula")
                
                # If we couldn't find the total rows, try using hardcoded values based on your template
                if not total_ht_row:
                    # Assume the totals are at fixed positions based on your template structure
                    try:
                        # Common positions for totals - adjust based on your template
                        subtotal_range = f"J{start_row}:J{start_row + len(invoice_items) - 1}"
                        
                        # Try to find cells that contain "total ht", "tva", etc.
                        found = False
                        for row in range(35, 45):
                            if found:
                                break
                            for col in range(5, 10):
                                cell_value = sheet.cell(row=row, column=col).value
                                if cell_value and isinstance(cell_value, str) and "total" in cell_value.lower():
                                    # Found a total row, assume it's the start of the totals section
                                    sheet.cell(row=row, column=10).value = f"=SUM({subtotal_range})"
                                    sheet.cell(row=row+1, column=10).value = f"=J{row}*0.2"
                                    sheet.cell(row=row+2, column=10).value = f"=J{row}+J{row+1}"
                                    found = True
                                    break
                    except:
                        print("Could not set totals with fallback method")
                
                # If this is not the first invoice, add note about it being a continuation
                if invoice_index > 0:
                    try:
                        note_cell = sheet.cell(row=start_row - 2, column=1)
                        note_cell.value = f"Suite de la facture {initial_display_id}"
                        note_cell.font = Font(bold=True)
                    except:
                        print("Could not set continuation note")
                    
                    try:
                        continuation_cell = sheet.cell(row=3, column=1)
                        continuation_cell.value = f"Facture {invoice_index + 1}/{num_invoices}"
                        continuation_cell.font = Font(bold=True)
                    except:
                        print("Could not set continuation header")
                        
                elif num_invoices > 1:
                    try:
                        multi_invoice_cell = sheet.cell(row=3, column=1)
                        multi_invoice_cell.value = f"Facture 1/{num_invoices}"
                        multi_invoice_cell.font = Font(bold=True)
                    except:
                        print("Could not set multi-invoice header")
                
                # Save the workbook
                try:
                    workbook.save(current_output_file)
                except Exception as e:
                    print(f"Error saving workbook: {str(e)}")
                    # If we can't save, try to create a new file with the data
                    self._write_fallback_file(current_output_file, invoice_items, current_display_id, client_info)
                
                # Add to the list of generated invoices
                generated_invoices.append(current_output_file)
            
            # Return the paths of all generated invoices
            if len(generated_invoices) == 1:
                return generated_invoices[0]  # For backward compatibility
            else:
                return generated_invoices
        
        except Exception as e:
            print(f"Error creating invoice: {str(e)}")
            raise
        
def _write_fallback_file(self, output_file, line_items, invoice_id, client_info=None):
    """Fallback method to write data to a new Excel file if we can't modify the template"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Set header information
    ws['A1'] = "FACTURE"
    ws['A1'].font = Font(bold=True, size=16)
    
    ws['A3'] = "Facture N°:"
    ws['B3'] = invoice_id
    ws['B3'].font = Font(bold=True)
    
    ws['A4'] = "Date:"
    ws['B4'] = datetime.now().strftime("%d/%m/%Y")
    
    # Set client info if provided
    if client_info:
        row = 6
        if 'name' in client_info:
            ws['A6'] = "Client:"
            ws['B6'] = client_info['name']
            row = 7
        if 'address' in client_info:
            ws[f'A{row}'] = "Adresse:"
            ws[f'B{row}'] = client_info['address']
            row += 1
        if 'ice' in client_info:
            ws[f'A{row}'] = "ICE:"
            ws[f'B{row}'] = client_info['ice']
    
    # Create header row for items
    row = 10
    headers = ["Description", "Quantité", "Prix unitaire", "Total"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=row, column=col).value = header
        ws.cell(row=row, column=col).font = Font(bold=True)
    
    # Add line items
    for idx, item in enumerate(line_items):
        row = 11 + idx
        # Description
        ws.cell(row=row, column=1).value = item['description']
        # Quantity
        ws.cell(row=row, column=2).value = round(item['quantity'], 2)
        # Unit price
        ws.cell(row=row, column=3).value = round(item['unit_price'], 2)
        # Total (formula)
        ws.cell(row=row, column=4).value = f"=B{row}*C{row}"
    
    # Add totals
    total_row = 11 + len(line_items) + 1
    ws.cell(row=total_row, column=3).value = "Total HT:"
    ws.cell(row=total_row, column=3).font = Font(bold=True)
    ws.cell(row=total_row, column=4).value = f"=SUM(D11:D{10 + len(line_items)})"
    
    ws.cell(row=total_row + 1, column=3).value = "TVA (20%):"
    ws.cell(row=total_row + 1, column=3).font = Font(bold=True)
    ws.cell(row=total_row + 1, column=4).value = f"=D{total_row}*0.2"
    
    ws.cell(row=total_row + 2, column=3).value = "Total TTC:"
    ws.cell(row=total_row + 2, column=3).font = Font(bold=True)
    ws.cell(row=total_row + 2, column=4).value = f"=D{total_row}+D{total_row + 1}"
    
    # Save this new workbook
    wb.save(output_file)
        
def _write_fallback_file(self, output_file, line_items, invoice_id, client_info=None):
    """Fallback method to write data to a new Excel file if we can't modify the template"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Set header information
    ws['A1'] = "FACTURE"
    ws['A1'].font = Font(bold=True, size=16)
    
    ws['A3'] = "Facture N°:"
    ws['B3'] = invoice_id
    ws['B3'].font = Font(bold=True)
    
    ws['A4'] = "Date:"
    ws['B4'] = datetime.now().strftime("%d/%m/%Y")
    
    # Set client info if provided
    if client_info:
        row = 6
        if 'name' in client_info:
            ws['A6'] = "Client:"
            ws['B6'] = client_info['name']
            row = 7
        if 'address' in client_info:
            ws[f'A{row}'] = "Adresse:"
            ws[f'B{row}'] = client_info['address']
            row += 1
        if 'ice' in client_info:
            ws[f'A{row}'] = "ICE:"
            ws[f'B{row}'] = client_info['ice']
    
    # Create header row for items
    row = 10
    headers = ["Description", "Quantité", "Prix unitaire", "Total"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=row, column=col).value = header
        ws.cell(row=row, column=col).font = Font(bold=True)
    
    # Add line items
    for idx, item in enumerate(line_items):
        row = 11 + idx
        # Description
        ws.cell(row=row, column=1).value = item['description']
        # Quantity
        ws.cell(row=row, column=2).value = round(item['quantity'], 2)
        # Unit price
        ws.cell(row=row, column=3).value = round(item['unit_price'], 2)
        # Total (formula)
        ws.cell(row=row, column=4).value = f"=B{row}*C{row}"
    
    # Add totals
    total_row = 11 + len(line_items) + 1
    ws.cell(row=total_row, column=3).value = "Total HT:"
    ws.cell(row=total_row, column=3).font = Font(bold=True)
    ws.cell(row=total_row, column=4).value = f"=SUM(D11:D{10 + len(line_items)})"
    
    ws.cell(row=total_row + 1, column=3).value = "TVA (20%):"
    ws.cell(row=total_row + 1, column=3).font = Font(bold=True)
    ws.cell(row=total_row + 1, column=4).value = f"=D{total_row}*0.2"
    
    ws.cell(row=total_row + 2, column=3).value = "Total TTC:"
    ws.cell(row=total_row + 2, column=3).font = Font(bold=True)
    ws.cell(row=total_row + 2, column=4).value = f"=D{total_row}+D{total_row + 1}"
    
    # Save this new workbook
    wb.save(output_file)
            
    def _write_fallback_file(self, output_file, line_items, invoice_id, client_info=None):
        """Fallback method to write data to a new Excel file if we can't modify the template"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Set header information
        ws['A1'] = "FACTURE"
        ws['A1'].font = Font(bold=True, size=16)
        
        ws['A3'] = "Facture N°:"
        ws['B3'] = invoice_id
        ws['B3'].font = Font(bold=True)
        
        ws['A4'] = "Date:"
        ws['B4'] = datetime.now().strftime("%d/%m/%Y")
        
        # Set client info if provided
        if client_info:
            row = 6
            if 'name' in client_info:
                ws['A6'] = "Client:"
                ws['B6'] = client_info['name']
                row = 7
            if 'address' in client_info:
                ws[f'A{row}'] = "Adresse:"
                ws[f'B{row}'] = client_info['address']
                row += 1
            if 'ice' in client_info:
                ws[f'A{row}'] = "ICE:"
                ws[f'B{row}'] = client_info['ice']
        
        # Create header row for items
        row = 10
        headers = ["Description", "Quantité", "Prix unitaire", "Total"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=row, column=col).value = header
            ws.cell(row=row, column=col).font = Font(bold=True)
        
        # Add line items
        for idx, item in enumerate(line_items):
            row = 11 + idx
            # Description
            ws.cell(row=row, column=1).value = item['description']
            # Quantity
            ws.cell(row=row, column=2).value = round(item['quantity'], 2)
            # Unit price
            ws.cell(row=row, column=3).value = round(item['unit_price'], 2)
            # Total (formula)
            ws.cell(row=row, column=4).value = f"=B{row}*C{row}"
        
        # Add totals
        total_row = 11 + len(line_items) + 1
        ws.cell(row=total_row, column=3).value = "Total HT:"
        ws.cell(row=total_row, column=3).font = Font(bold=True)
        ws.cell(row=total_row, column=4).value = f"=SUM(D11:D{10 + len(line_items)})"
        
        ws.cell(row=total_row + 1, column=3).value = "TVA (20%):"
        ws.cell(row=total_row + 1, column=3).font = Font(bold=True)
        ws.cell(row=total_row + 1, column=4).value = f"=D{total_row}*0.2"
        
        ws.cell(row=total_row + 2, column=3).value = "Total TTC:"
        ws.cell(row=total_row + 2, column=3).font = Font(bold=True)
        ws.cell(row=total_row + 2, column=4).value = f"=D{total_row}+D{total_row + 1}"
        
        # Save this new workbook
        wb.save(output_file)

class InvoiceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Générateur de Factures")
        self.root.geometry("750x600")
        self.root.configure(padx=20, pady=20)
        # Initialize the invoice generator
        self.generator = InvoiceGenerator()
        
        # Check if template file exists
        if not os.path.exists(self.generator.template_path):
            messagebox.showwarning(
                "Fichier modèle manquant",
                f"Le fichier modèle n'a pas été trouvé à l'emplacement:\n{self.generator.template_path}\n\n"
                "Veuillez sélectionner le fichier modèle lors de la première utilisation."
            )
        
        # Default data file path
        self.data_file = os.path.join(self.generator.base_path, "AZZOUZIFCT.xlsx")
        
        # Check if data file exists
        if not os.path.exists(self.data_file):
            # Set to empty and will require user selection
            self.data_file = ""
        
        # Create the main frame
        self.main_frame = ttk.Frame(root)
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Create styles
        self.style = ttk.Style()
        self.style.configure("TLabel", font=("Arial", 11))
        self.style.configure("TButton", font=("Arial", 11))
        self.style.configure("TEntry", font=("Arial", 11))
        self.style.configure("Header.TLabel", font=("Arial", 14, "bold"))
        
        # Create the header
        header = ttk.Label(self.main_frame, text="Générateur de Factures", style="Header.TLabel")
        header.pack(pady=(0, 20))
        
        # Create the form frame
        form_frame = ttk.Frame(self.main_frame)
        form_frame.pack(fill=tk.X, pady=10)
        
        # Data file selection
        file_frame = ttk.Frame(form_frame)
        file_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(file_frame, text="Fichier de données:").pack(side=tk.LEFT, padx=(0, 10))
        
        self.data_file_var = tk.StringVar(value=self.data_file)
        data_file_entry = ttk.Entry(file_frame, textvariable=self.data_file_var, width=40)
        data_file_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        browse_btn = ttk.Button(file_frame, text="Parcourir...", command=self.browse_data_file)
        browse_btn.pack(side=tk.LEFT)
        
        # Invoice ID
        id_frame = ttk.Frame(form_frame)
        id_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(id_frame, text="Numéro de facture:").pack(side=tk.LEFT, padx=(0, 10))
        
        self.invoice_id_var = tk.StringVar()
        invoice_id_entry = ttk.Entry(id_frame, textvariable=self.invoice_id_var, width=15)
        invoice_id_entry.pack(side=tk.LEFT)
        
        ttk.Label(id_frame, text="(Laissez vide pour générer automatiquement)").pack(side=tk.LEFT, padx=(10, 0))
        
        # Client info
        client_frame = ttk.LabelFrame(form_frame, text="Informations client")
        client_frame.pack(fill=tk.X, pady=10)
        
        # Client name
        name_frame = ttk.Frame(client_frame)
        name_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(name_frame, text="Nom / Société:").pack(side=tk.LEFT, padx=(0, 10))
        
        self.client_name_var = tk.StringVar()
        client_name_entry = ttk.Entry(name_frame, textvariable=self.client_name_var, width=40)
        client_name_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Client address
        address_frame = ttk.Frame(client_frame)
        address_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(address_frame, text="Adresse:").pack(side=tk.LEFT, padx=(0, 10))
        
        self.client_address_var = tk.StringVar()
        client_address_entry = ttk.Entry(address_frame, textvariable=self.client_address_var, width=40)
        client_address_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Client ICE
        ice_frame = ttk.Frame(client_frame)
        ice_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(ice_frame, text="ICE:").pack(side=tk.LEFT, padx=(0, 10))
        
        self.client_ice_var = tk.StringVar()
        client_ice_entry = ttk.Entry(ice_frame, textvariable=self.client_ice_var, width=40)
        client_ice_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Output options
        output_frame = ttk.LabelFrame(form_frame, text="Options de sortie")
        output_frame.pack(fill=tk.X, pady=10)
        
        # Output folder
        folder_frame = ttk.Frame(output_frame)
        folder_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(folder_frame, text="Dossier de sortie:").pack(side=tk.LEFT, padx=(0, 10))
        
        self.output_folder_var = tk.StringVar(value=self.generator.output_folder)
        output_folder_entry = ttk.Entry(folder_frame, textvariable=self.output_folder_var, width=40)
        output_folder_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        browse_output_btn = ttk.Button(folder_frame, text="Parcourir...", command=self.browse_output_folder)
        browse_output_btn.pack(side=tk.LEFT)
        
        # Buttons
        button_frame = ttk.Frame(self.main_frame)
        button_frame.pack(fill=tk.X, pady=20)
        
        self.generate_btn = ttk.Button(button_frame, text="Générer la facture", command=self.generate_invoice)
        self.generate_btn.pack(side=tk.RIGHT, padx=5)
        
        reset_btn = ttk.Button(button_frame, text="Réinitialiser", command=self.reset_form)
        reset_btn.pack(side=tk.RIGHT, padx=5)
        
        # Log area
        log_frame = ttk.LabelFrame(self.main_frame, text="Journal")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        self.log_text = ScrolledText(log_frame, height=10)
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.log_text.config(state=tk.DISABLED)
        
        # Status bar
        self.status_var = tk.StringVar(value="Prêt")
        status_bar = ttk.Label(self.main_frame, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(fill=tk.X, side=tk.BOTTOM, pady=(10, 0))
        
    def log(self, message):
        """Add a message to the log area"""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, f"{datetime.now().strftime('%H:%M:%S')} - {message}\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        
    def browse_data_file(self):
        """Open file dialog to select data file"""
        file_path = filedialog.askopenfilename(
            title="Sélectionner le fichier de données",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
            initialdir=self.generator.base_path
        )
        
        if file_path:
            self.data_file_var.set(file_path)
            self.log(f"Fichier de données sélectionné: {file_path}")
    
    def browse_output_folder(self):
        """Open folder dialog to select output folder"""
        folder_path = filedialog.askdirectory(
            title="Sélectionner le dossier de sortie",
            initialdir=self.generator.output_folder
        )
        
        if folder_path:
            self.output_folder_var.set(folder_path)
            self.log(f"Dossier de sortie sélectionné: {folder_path}")
    
    def generate_invoice(self):
        """Generate the invoice(s) with the provided information"""
        try:
            # Disable the generate button during processing
            self.generate_btn.config(state=tk.DISABLED)
            self.status_var.set("Génération en cours...")
            self.root.update()
            
            # Get values from form
            data_file = self.data_file_var.get()
            invoice_id = self.invoice_id_var.get() if self.invoice_id_var.get() else None
            output_folder = self.output_folder_var.get()
            
            # Create client info dictionary
            client_info = {}
            if self.client_name_var.get():
                client_info['name'] = self.client_name_var.get()
            if self.client_address_var.get():
                client_info['address'] = self.client_address_var.get()
            if self.client_ice_var.get():
                client_info['ice'] = self.client_ice_var.get()
            
            # Validate inputs
            if not os.path.exists(data_file):
                raise ValueError(f"Le fichier de données n'existe pas: {data_file}")
                
            if not os.path.exists(self.generator.template_path):
                raise ValueError(f"Le fichier modèle n'existe pas: {self.generator.template_path}")
                
            if not os.path.exists(output_folder):
                # Create output folder if it doesn't exist
                try:
                    os.makedirs(output_folder, exist_ok=True)
                    self.log(f"Dossier de sortie créé: {output_folder}")
                except Exception as e:
                    raise ValueError(f"Impossible de créer le dossier de sortie: {str(e)}")
            
            # Update generator paths
            self.generator.output_folder = output_folder
            
            # Log start of generation
            self.log(f"Début de la génération de la facture...")
            self.log(f"Fichier de données: {data_file}")
            self.log(f"Fichier modèle: {self.generator.template_path}")
            
            # Generate filename if invoice_id provided
            output_file = None
            if invoice_id:
                output_file = os.path.join(output_folder, f"invoice_{invoice_id}.xlsx")
            
            # Generate the invoice(s)
            result = self.generator.create_invoice(
                data_file,
                invoice_id,
                output_file,
                client_info
            )
            
            # Handle the result (could be a single path or a list of paths)
            if isinstance(result, list):
                # Multiple invoices were generated
                self.log(f"{len(result)} factures générées avec succès:")
                for idx, invoice_path in enumerate(result):
                    self.log(f"  {idx+1}. {invoice_path}")
                
                self.status_var.set(f"{len(result)} factures générées")
                
                # Show success message
                messagebox.showinfo(
                    "Génération réussie", 
                    f"{len(result)} factures ont été générées avec succès.\n\nDossier: {output_folder}"
                )
                
                # Ask if user wants to open the output folder
                if messagebox.askyesno("Ouvrir le dossier", "Voulez-vous ouvrir le dossier contenant les factures générées?"):
                    import platform
                    import subprocess
                    
                    system = platform.system()
                    try:
                        if system == 'Windows':
                            os.startfile(output_folder)
                        elif system == 'Darwin':  # macOS
                            subprocess.call(['open', output_folder])
                        else:  # Linux and other Unix-like systems
                            subprocess.call(['xdg-open', output_folder])
                    except Exception as e:
                        self.log(f"Erreur lors de l'ouverture du dossier: {str(e)}")
                        messagebox.showwarning("Avertissement", f"Impossible d'ouvrir le dossier automatiquement.\n\nLes factures ont été enregistrées ici:\n{output_folder}")
            else:
                # Single invoice was generated
                self.log(f"Facture générée avec succès: {result}")
                self.status_var.set(f"Facture générée: {os.path.basename(result)}")
                
                # Show success message
                messagebox.showinfo(
                    "Génération réussie", 
                    f"La facture a été générée avec succès.\n\nFichier: {result}"
                )
                
                # Ask if user wants to open the generated file
                if messagebox.askyesno("Ouvrir le fichier", "Voulez-vous ouvrir la facture générée?"):
                    import platform
                    import subprocess
                    
                    system = platform.system()
                    try:
                        if system == 'Windows':
                            os.startfile(result)
                        elif system == 'Darwin':  # macOS
                            subprocess.call(['open', result])
                        else:  # Linux and other Unix-like systems
                            subprocess.call(['xdg-open', result])
                    except Exception as e:
                        self.log(f"Erreur lors de l'ouverture du fichier: {str(e)}")
                        messagebox.showwarning("Avertissement", f"Impossible d'ouvrir le fichier automatiquement.\n\nLe fichier a été enregistré ici:\n{result}")
        
        except Exception as e:
            # Log error
            error_msg = str(e)
            self.log(f"Erreur: {error_msg}")
            self.status_var.set("Erreur lors de la génération")
            
            # Show error message
            messagebox.showerror("Erreur", f"Une erreur s'est produite lors de la génération de la facture:\n\n{error_msg}")
        
        finally:
            # Re-enable the generate button
            self.generate_btn.config(state=tk.NORMAL)
    
    def reset_form(self):
        """Reset the form to default values"""
        self.invoice_id_var.set("")
        self.client_name_var.set("")
        self.client_address_var.set("")
        self.client_ice_var.set("")
        self.data_file_var.set(os.path.join(self.generator.base_path, "AZZOUZIFCT.xlsx"))
        self.output_folder_var.set(self.generator.base_path)
        
        self.log("Formulaire réinitialisé")
        self.status_var.set("Prêt")


def add_menu(root, app):
    """Add menu bar to the application"""
    menubar = tk.Menu(root)
    
    # File menu
    filemenu = tk.Menu(menubar, tearoff=0)
    filemenu.add_command(label="Sélectionner fichier de données", command=app.browse_data_file)
    filemenu.add_command(label="Sélectionner dossier de sortie", command=app.browse_output_folder)
    filemenu.add_separator()
    filemenu.add_command(label="Quitter", command=root.quit)
    menubar.add_cascade(label="Fichier", menu=filemenu)
    
    # Template menu
    templatemenu = tk.Menu(menubar, tearoff=0)
    templatemenu.add_command(label="Sélectionner modèle", command=lambda: select_template(app))
    menubar.add_cascade(label="Modèle", menu=templatemenu)
    
    # Help menu
    helpmenu = tk.Menu(menubar, tearoff=0)
    helpmenu.add_command(label="À propos", command=lambda: show_about(root))
    menubar.add_cascade(label="Aide", menu=helpmenu)
    
    root.config(menu=menubar)

def select_template(app):
    """Allow user to select a template file"""
    file_path = filedialog.askopenfilename(
        title="Sélectionner le fichier modèle",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        initialdir=app.generator.base_path
    )
    
    if file_path:
        app.generator.template_path = file_path
        app.log(f"Fichier modèle sélectionné: {file_path}")

def show_about(root):
    """Show about dialog"""
    messagebox.showinfo(
        "À propos",
        "Générateur de Factures\n\n"
        "Version 1.0\n\n"
        "Un programme pour générer des factures à partir de fichiers Excel.\n"
        "Compatible avec Windows et Linux."
    )

if __name__ == "__main__":
    # Handle high DPI displays on Windows
    if os.name == 'nt':
        try:
            from ctypes import windll
            windll.shcore.SetProcessDpiAwareness(1)
        except:
            pass
    
    root = tk.Tk()
    app = InvoiceApp(root)
    add_menu(root, app)
    root.mainloop()